"""
Экспорт отзывов и вопросов в .xlsx + создание zip-архива.

Имена файлов содержат timestamp, чтобы повторный сбор того же артикула
с тем же фильтром не перезаписывал файлы из предыдущих сборов
(иначе история бы ссылалась на устаревшие данные).
"""
import os
import re
import zipfile
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import FILES_DIR


# ─── Стили ───────────────────────────────────────────────────────────────────

def _header_style(ws, headers: list[str], fill_color: str = "1F4E79") -> None:
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        bottom=Side(style="medium", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
    )
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def _auto_width(ws, min_width: int = 10, max_width: int = 60) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                # считаем максимальную длину первой строки (для многострочных ячеек)
                first_line = str(cell.value).split("\n", 1)[0]
                if len(first_line) > max_len:
                    max_len = len(first_line)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _row_style(ws, row_num: int, even: bool) -> None:
    fill_color = "DCE6F1" if even else "FFFFFF"
    fill = PatternFill("solid", fgColor=fill_color)
    for cell in ws[row_num]:
        cell.fill = fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)


# ─── Безопасные имена файлов ─────────────────────────────────────────────────

_UNSAFE = re.compile(r"[^\w\-а-яА-ЯёЁ]+", re.UNICODE)


def _safe_label(label: str) -> str:
    """Убрать эмодзи/пробелы/символы, оставить буквы/цифры/дефисы."""
    label = label.replace("–", "-").replace(" ", "_")
    label = _UNSAFE.sub("", label)
    return label.strip("_-") or "все"


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


# ─── Отзывы ──────────────────────────────────────────────────────────────────

def save_reviews_xlsx(reviews: list[dict], article: str, filter_label: str) -> str:
    os.makedirs(FILES_DIR, exist_ok=True)

    safe_label = _safe_label(filter_label)
    filename = f"Отзывы_{article}_{safe_label}_{_timestamp()}.xlsx"
    filepath = os.path.join(FILES_DIR, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отзывы"

    headers = [
        "Дата", "Автор", "Оценка", "Текст",
        "Достоинства", "Недостатки",
        "Ответ продавца", "Дата ответа", "Фото/видео",
    ]
    _header_style(ws, headers, fill_color="1F4E79")
    ws.row_dimensions[1].height = 30

    for i, review in enumerate(reviews, 2):
        ws.append([
            review.get("Дата", ""),
            review.get("Автор", ""),
            review.get("Оценка", ""),
            review.get("Текст", ""),
            review.get("Достоинства", ""),
            review.get("Недостатки", ""),
            review.get("Ответ продавца", ""),
            review.get("Дата ответа", ""),
            review.get("Фото/видео", 0),
        ])
        _row_style(ws, i, even=(i % 2 == 0))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)
    wb.save(filepath)
    return filepath


# ─── Вопросы ─────────────────────────────────────────────────────────────────

def save_questions_xlsx(questions: list[dict], article: str) -> str:
    os.makedirs(FILES_DIR, exist_ok=True)

    filename = f"Вопросы_{article}_{_timestamp()}.xlsx"
    filepath = os.path.join(FILES_DIR, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Вопросы"

    headers = [
        "Дата", "Автор", "Вопрос",
        "Ответ продавца", "Дата ответа продавца", "Ответы покупателей",
    ]
    _header_style(ws, headers, fill_color="375623")
    ws.row_dimensions[1].height = 30

    for i, q in enumerate(questions, 2):
        ws.append([
            q.get("Дата", ""),
            q.get("Автор", ""),
            q.get("Вопрос", ""),
            q.get("Ответ продавца", ""),
            q.get("Дата ответа продавца", ""),
            q.get("Ответы покупателей", ""),
        ])
        _row_style(ws, i, even=(i % 2 == 0))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)
    wb.save(filepath)
    return filepath


# ─── ZIP-архив ───────────────────────────────────────────────────────────────

def create_archive(reviews_path: str, questions_path: str, article: str) -> str:
    os.makedirs(FILES_DIR, exist_ok=True)
    archive_path = os.path.join(
        FILES_DIR, f"WB_{article}_{_timestamp()}.zip"
    )

    with zipfile.ZipFile(archive_path, "w", zipfile.ZIP_DEFLATED) as zf:
        if reviews_path and os.path.exists(reviews_path):
            zf.write(reviews_path, os.path.basename(reviews_path))
        if questions_path and os.path.exists(questions_path):
            zf.write(questions_path, os.path.basename(questions_path))

    return archive_path
